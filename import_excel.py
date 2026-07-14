"""One-time import: convert Carlos's Excel report to JSONL format."""
import json
import os
import sys
from datetime import datetime

import openpyxl

import config
import order_store


def import_closer_sheet(filepath: str):
    """Import the 'Closer' sheet from Carlos's Excel into orders_log.jsonl."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Closer"]

    records = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        tienda = str(row[0] or "").strip()
        pedido = str(row[1] or "").strip()
        fecha = row[2]
        observaciones = str(row[5] or "").strip()
        tipo_venta = str(row[6] or "").strip()
        importe = row[7]

        if not pedido or not pedido.startswith("#"):
            continue

        order_number_str = pedido.replace("#", "").strip()
        try:
            order_number_int = int(order_number_str)
        except ValueError:
            continue

        if isinstance(fecha, datetime):
            created_at = fecha.isoformat()
        else:
            created_at = str(fecha) if fecha else ""

        if isinstance(importe, (int, float)):
            total_price = f"{importe:.2f}"
        else:
            total_price = "0.00"

        payment_type = "cod" if "cash" in tipo_venta.lower() or "delivery" in tipo_venta.lower() else "paid"

        record = {
            "order_id": order_number_int,
            "order_number": pedido,
            "created_at": created_at,
            "customer_name": "N/A",
            "landing_site": "",
            "utm_source": "carlos_excel_import",
            "utm_medium": None,
            "payment_type": payment_type,
            "payment_gateway": tipo_venta,
            "financial_status": "paid" if payment_type == "paid" else "pending",
            "total_price": total_price,
            "currency": config.CURRENCY,
            "line_items": [observaciones] if observaciones else [],
            "fulfillment_status": None,
            "shipment_status": None,
            "tracking_company": None,
            "tracking_number": None,
            "tracking_url": None,
            "cod_verified": False,
            "source": "excel_import",
        }
        records.append(record)

    order_store.append_orders(records)
    print(f"Imported {len(records)} orders from Excel")

    paid = [r for r in records if r["payment_type"] == "paid"]
    cod = [r for r in records if r["payment_type"] == "cod"]
    print(f"  Paid: {len(paid)} (EUR {sum(float(r['total_price']) for r in paid):,.2f})")
    print(f"  COD:  {len(cod)} (EUR {sum(float(r['total_price']) for r in cod):,.2f})")


def import_hours_sheet(filepath: str):
    """Import the 'Horas' sheet and print summary."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Horas"]

    print("\n=== Horas de trabajo ===")
    days = 0
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        fecha = row[1]
        inicio = row[2]
        fin = row[3]
        total_tiempo = row[6]

        if not fecha:
            continue
        days += 1
        fecha_str = fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else str(fecha)
        print(f"  {fecha_str}: {inicio} - {fin} ({total_tiempo})")

    print(f"\nTotal dias trabajados: {days}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = os.path.join(os.path.dirname(__file__), "Carlos (06-2026).xlsx")

    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)

    print(f"Importing from: {filepath}\n")
    import_closer_sheet(filepath)
    import_hours_sheet(filepath)
